"""Editable salary view — writes changes back to the source Excel file."""

import openpyxl
import pandas as pd
import streamlit as st
from pathlib import Path
from dashboard.components.charts import money
from dashboard.components.owner_pl_tree import _month_key

# Column mapping for the "Отчет по ЗП" sheet (1-based, see parser/margin_report.py:138)
_COL_MAP = {
    "group":            1,
    "role":             2,
    "name":             3,
    "fiks":             4,
    "manager_bonus":    7,
    "specialist_bonus": 8,
    "activity":         9,
    "other_pay":       11,
    "vacation_sick":   12,
    "total_accrued":   13,
    "paid_1c":         14,
}
_DATA_START_ROW = 5
_NAME_PREFIX = "ФИО"


def _save_salary_to_excel(month: str, df_edited: pd.DataFrame) -> tuple[bool, str]:
    from parser.margin_report import FILES
    if month not in FILES:
        return False, f"Файл за {month} не найден"
    path: Path = FILES[month]

    try:
        wb = openpyxl.load_workbook(path)
    except Exception as e:
        return False, f"Не удалось открыть {path.name}: {e}"

    if "Отчет по ЗП" not in wb.sheetnames:
        return False, f"В файле {path.name} нет листа «Отчет по ЗП»"

    ws = wb["Отчет по ЗП"]

    clean = df_edited.copy()
    clean["name"] = clean["name"].fillna("").astype(str).str.strip()
    clean = clean[clean["name"] != ""]

    # vacation_other → split: vacation_sick gets full value, other_pay = 0
    if "vacation_other" in clean.columns:
        clean["vacation_sick"] = pd.to_numeric(clean["vacation_other"], errors="coerce").fillna(0.0)
        clean["other_pay"] = 0.0
    if "paid_1c" not in clean.columns:
        clean["paid_1c"] = 0.0

    component_cols = ["fiks", "manager_bonus", "specialist_bonus", "activity",
                      "other_pay", "vacation_sick"]
    for c in component_cols:
        if c not in clean.columns:
            clean[c] = 0.0
        clean[c] = pd.to_numeric(clean[c], errors="coerce").fillna(0.0)
    clean["total_accrued"] = clean[component_cols].sum(axis=1)

    last_row = max(ws.max_row, _DATA_START_ROW)
    for r in range(_DATA_START_ROW, last_row + 1):
        for col in _COL_MAP.values():
            ws.cell(row=r, column=col).value = None

    for offset, (_, rec) in enumerate(clean.iterrows()):
        row = _DATA_START_ROW + offset
        nm = str(rec["name"]).strip()
        if not nm.startswith(_NAME_PREFIX):
            nm = f"{_NAME_PREFIX} {nm}"
        ws.cell(row=row, column=_COL_MAP["group"]).value = rec.get("group", "") or ""
        ws.cell(row=row, column=_COL_MAP["role"]).value  = rec.get("role", "")  or ""
        ws.cell(row=row, column=_COL_MAP["name"]).value  = nm
        ws.cell(row=row, column=_COL_MAP["fiks"]).value             = float(rec.get("fiks", 0) or 0)
        ws.cell(row=row, column=_COL_MAP["manager_bonus"]).value    = float(rec.get("manager_bonus", 0) or 0)
        ws.cell(row=row, column=_COL_MAP["specialist_bonus"]).value = float(rec.get("specialist_bonus", 0) or 0)
        ws.cell(row=row, column=_COL_MAP["activity"]).value         = float(rec.get("activity", 0) or 0)
        ws.cell(row=row, column=_COL_MAP["other_pay"]).value        = float(rec.get("other_pay", 0) or 0)
        ws.cell(row=row, column=_COL_MAP["vacation_sick"]).value    = float(rec.get("vacation_sick", 0) or 0)
        ws.cell(row=row, column=_COL_MAP["total_accrued"]).value    = float(rec.get("total_accrued", 0) or 0)
        ws.cell(row=row, column=_COL_MAP["paid_1c"]).value          = 0.0

    try:
        wb.save(path)
    except PermissionError:
        return False, f"Файл {path.name} открыт в Excel — закройте и попробуйте снова."
    except Exception as e:
        return False, f"Ошибка сохранения: {e}"
    return True, f"Сохранено: {path.name} ({len(clean)} сотрудников)"


def render_salary_editor(salary_df: pd.DataFrame, months: list[str]) -> None:
    st.subheader("👥 Зарплаты сотрудников — редактирование")
    st.caption(
        "Изменения сохраняются прямо в исходный Excel-файл за выбранный месяц. "
        "«Итого начислено ★» рассчитывается автоматически как сумма всех компонентов."
    )

    if salary_df is None or salary_df.empty:
        st.info("Нет данных по зарплатам")
        return

    available = sorted(
        [m for m in months if (salary_df["month"] == m).any()],
        key=_month_key,
    )
    if not available:
        st.info("Нет данных по зарплате за выбранные месяцы")
        return

    month = st.selectbox("Месяц для редактирования", available,
                         index=len(available) - 1, key="salary_edit_month")
    df_m = salary_df[salary_df["month"] == month].copy()
    df_m = df_m.drop(columns=["month"], errors="ignore")

    if "name" in df_m.columns:
        df_m["name"] = df_m["name"].astype(str).str.replace(r"^ФИО\s+", "", regex=True)

    # Merge vacation_sick + other_pay → vacation_other (п. 11)
    df_m["vacation_other"] = (
        pd.to_numeric(df_m.get("vacation_sick", 0), errors="coerce").fillna(0.0)
        + pd.to_numeric(df_m.get("other_pay", 0), errors="coerce").fillna(0.0)
    )

    # Column order: total_accrued first after identity cols (п. 9)
    edit_cols_order = ["name", "role", "group", "total_accrued", "fiks",
                       "manager_bonus", "specialist_bonus", "activity", "vacation_other"]
    df_m = df_m.reindex(columns=[c for c in edit_cols_order if c in df_m.columns])

    edited = st.data_editor(
        df_m,
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "name":             st.column_config.TextColumn("ФИО"),
            "role":             st.column_config.TextColumn("Роль / должность"),
            "group":            st.column_config.TextColumn("Группа"),
            "total_accrued":    st.column_config.NumberColumn("Итого начислено ★", format="%.0f",
                                                              disabled=True,
                                                              help="Пересчитывается при сохранении."),
            "fiks":             st.column_config.NumberColumn("ФИКС",        format="%.0f", step=1000.0),
            "manager_bonus":    st.column_config.NumberColumn("Бонус M",     format="%.0f", step=1000.0),
            "specialist_bonus": st.column_config.NumberColumn("Бонус S",     format="%.0f", step=1000.0),
            "activity":         st.column_config.NumberColumn("KPI",         format="%.0f", step=1000.0),
            "vacation_other":   st.column_config.NumberColumn("Отпуск и пр.", format="%.0f", step=1000.0),
        },
        key=f"salary_editor_{month}",
    )

    # п. 3: убрана кнопка "Сбросить"
    save_clicked = st.button("💾 Сохранить в Excel", type="primary",
                             key=f"salary_save_{month}")

    if save_clicked:
        ok, msg = _save_salary_to_excel(month, edited)
        if ok:
            st.success(msg)
            st.cache_data.clear()
            st.rerun()
        else:
            st.error(msg)

    # ── Summary footer ────────────────────────────────────────────────────
    st.divider()
    num_cols = ["fiks", "manager_bonus", "specialist_bonus", "activity", "vacation_other"]
    edited_num = edited.copy()
    for c in num_cols:
        if c in edited_num.columns:
            edited_num[c] = pd.to_numeric(edited_num[c], errors="coerce").fillna(0.0)
    total_now = edited_num[[c for c in num_cols if c in edited_num.columns]].sum().sum()
    headcount = len([n for n in edited["name"].fillna("") if str(n).strip()])

    # п. 10: дельта к предыдущему месяцу
    prev_idx = available.index(month) - 1
    prev_month = available[prev_idx] if prev_idx >= 0 else None
    avg_now = total_now / headcount if headcount else 0.0

    def _delta_pct(curr, prev):
        if prev is None or prev == 0:
            return None
        return f"{(curr - prev) / abs(prev) * 100:+.1f}%".replace(".", ",")

    prev_total = prev_avg = prev_fiks_avg = prev_kpi_avg = None
    if prev_month is not None:
        prev_df = salary_df[salary_df["month"] == prev_month]
        if not prev_df.empty:
            prev_comp = ["fiks", "manager_bonus", "specialist_bonus",
                         "activity", "other_pay", "vacation_sick"]
            prev_total = float(
                prev_df[[c for c in prev_comp if c in prev_df.columns]].apply(
                    pd.to_numeric, errors="coerce"
                ).fillna(0).sum().sum()
            )
            prev_hc = prev_df["name"].nunique() or 1
            prev_avg = prev_total / prev_hc
            prev_fiks_avg = float(prev_df["fiks"].apply(pd.to_numeric, errors="coerce").fillna(0).sum()) / prev_hc
            prev_kpi_avg  = float(prev_df["activity"].apply(pd.to_numeric, errors="coerce").fillna(0).sum()) / prev_hc

    fiks_sum = float(edited_num["fiks"].sum()) if "fiks" in edited_num else 0.0
    kpi_sum  = float(edited_num["activity"].sum()) if "activity" in edited_num else 0.0
    fiks_avg = fiks_sum / headcount if headcount else 0.0
    kpi_avg  = kpi_sum  / headcount if headcount else 0.0

    s_cols = st.columns(4)
    s_cols[0].metric("Сотрудников", headcount)
    s_cols[1].metric("ФИКС (сумма)", money(fiks_sum))
    s_cols[2].metric("KPI (сумма)",  money(kpi_sum))
    s_cols[3].metric("Итого начислено", money(total_now),
                     delta=_delta_pct(total_now, prev_total))

    # п. 5: средние показатели
    st.caption("Средние на сотрудника:")
    a_cols = st.columns(3)
    a_cols[0].metric("Средняя ЗП",    money(avg_now),   delta=_delta_pct(avg_now,   prev_avg))
    a_cols[1].metric("Средний ФИКС",  money(fiks_avg),  delta=_delta_pct(fiks_avg,  prev_fiks_avg))
    a_cols[2].metric("Средний KPI",   money(kpi_avg),   delta=_delta_pct(kpi_avg,   prev_kpi_avg))
