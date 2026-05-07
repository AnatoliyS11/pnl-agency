"""
Parser for the optional 'Возвраты МП' sheet inside each monthly margin file.

Sheet schema (recommended):
  Row 1: headers — month | platform | client | refund_type | amount
  Row 2+: data rows

If the sheet is missing or empty, parse_refunds() returns an empty DataFrame.
This lets the dashboard show a placeholder until refund data is added.
"""

import openpyxl
import pandas as pd
from pathlib import Path

SHEET_NAME = "Возвраты МП"

# Map various header spellings to canonical column names
_HEADER_ALIASES = {
    "month":        ["month", "месяц"],
    "platform":     ["platform", "площадка", "маркетплейс", "мп"],
    "client":       ["client", "клиент", "проект", "project"],
    "refund_type":  ["refund_type", "тип возврата", "тип", "type"],
    "amount":       ["amount", "сумма", "руб", "amount_rub"],
}


def _normalize_headers(header_row: list) -> dict[int, str]:
    """Map column index → canonical name for known headers."""
    result = {}
    for idx, val in enumerate(header_row, start=1):
        if val is None:
            continue
        norm = str(val).strip().lower()
        for canon, aliases in _HEADER_ALIASES.items():
            if norm in aliases:
                result[idx] = canon
                break
    return result


def parse_refunds(path: Path, default_month: str | None = None) -> pd.DataFrame:
    """Parse the 'Возвраты МП' sheet from a single monthly file.

    `default_month` is used to fill the `month` column if the sheet doesn't have one.
    """
    if not Path(path).exists():
        return pd.DataFrame()

    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception:
        return pd.DataFrame()

    if SHEET_NAME not in wb.sheetnames:
        return pd.DataFrame()

    ws = wb[SHEET_NAME]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return pd.DataFrame()

    col_map = _normalize_headers(list(header_row))
    if "amount" not in col_map.values():
        return pd.DataFrame()

    data = []
    for raw in rows_iter:
        rec = {}
        for idx, canon in col_map.items():
            v = raw[idx - 1] if idx - 1 < len(raw) else None
            rec[canon] = v
        if rec.get("amount") in (None, "", 0):
            continue
        try:
            rec["amount"] = float(rec["amount"])
        except (TypeError, ValueError):
            continue
        rec.setdefault("month", default_month or "")
        rec["platform"]    = str(rec.get("platform")    or "Прочее")
        rec["client"]      = str(rec.get("client")      or "—")
        rec["refund_type"] = str(rec.get("refund_type") or "—")
        if not rec["month"]:
            rec["month"] = default_month or ""
        data.append(rec)

    return pd.DataFrame(data,
                        columns=["month", "platform", "client", "refund_type", "amount"])


def load_all_refunds() -> pd.DataFrame:
    """Aggregate refunds from every discovered monthly file."""
    from parser.margin_report import FILES  # avoid circular import at module load

    frames = []
    for month_label, path in FILES.items():
        df = parse_refunds(path, default_month=month_label)
        if not df.empty:
            frames.append(df)

    if not frames:
        return pd.DataFrame(columns=["month", "platform", "client", "refund_type", "amount"])
    return pd.concat(frames, ignore_index=True)
