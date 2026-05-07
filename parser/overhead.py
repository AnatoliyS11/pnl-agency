"""
Parser for overhead expenses file: Для ИИ V2 Расчет накладных расходов для PL.xlsx

Sheet "Расходы Шелковый путь":
  Row 1: month group headers (merged cells, e.g. "Январь 2026", "Февраль 2026", …)
  Row 2: sub-headers          (План / Прогноз / Факт)
  Row 3+: data rows
  Col A: category / subcategory name

Колонки месяцев обнаруживаются автоматически по merged-ячейкам строки 1
и подзаголовкам строки 2 — новый месяц появится сам, как только в Excel
добавят новый столбец-группу.

Strategy:
  - "Parent" category rows have at least one non-zero value in month columns.
  - Sub-rows (ФИКС, KPI, НДФЛ, etc.) have None values — they are breakdowns
    already aggregated in parent rows; we skip them.
  - "Общие расходы Юнита по Тарифу" rows are subtotals — skip to avoid double-counting.
  - ФОТ category rows (ФОТ Юнита, ФОТ производство, ФОТ руководители) are kept
    as placeholders (may be 0 when salary ledger not connected).
"""

import re
import openpyxl
import pandas as pd
from pathlib import Path

DATA_DIR = Path(__file__).parent.parent
OVERHEAD_FILE = DATA_DIR / "Для ИИ V2 Расчет накладных расходов для PL.xlsx"

_SKIP_NAMES = {
    # ── Subtotal markers (repeat their parent value, skip to avoid double-count) ──
    "Общие расходы Юнита по Тарифу",
    "Прочие расходы на персонал",   # = HR Юнита subtotal
    "Сервисы",                       # = Сервисы производство leaf (same value)
    # ── Parent/aggregate rows ──
    "Операционные расходы",          # = сумма Офис + Рекрутинг + Бизнес + IT
    "Отдел развития",                # = parent of Отдел продаж
    # ── Grand totals ──
    "Всего расходы  Юнита",
    "расходы без фот",
    # ── Financial operations (shown separately) ──
    "Финансовые операции",
    "Finance income / Финансовые доходы",
    "Finance outcome / Финансовые затраты",  # parent of Банковские услуги
    # ── Sub-items (None-valued breakdown rows) ──
    "ФОТ", "ФИКС", "KPI", "НДФЛ", "Страховые взносы", "Премиальный фонд 3%",
    "Аренда", "Содержание офиса", "Обустройство офиса", "Мебель",
    "Канц.товары", "Вода/еда", "Прочие", "Прочее",
    "Командировочные расходы", "Обучение", "Представительские расходы",
    "Внутренние мероприятия", "Управленческие расходы",
    "Аутсорсинг", "Курьер/Почта", "Связь/интернет", "Техника",
    "Прочие расходы", "Расходы на рабочее место",
    "Расходы на маркетинг/мероприятия",
}

# Category groups for grouping in the dashboard
_CATEGORY_GROUPS = {
    "ФОТ Юнита": "ФОТ (производство)",
    "ФОТ производство": "ФОТ (производство)",
    "ФОТ руководители Юнита": "ФОТ (руководство)",
    "Сервисы производство": "Сервисы",
    "HR Юнита": "Персонал",
    "HR": "Персонал",
    "Рекрутинг": "Персонал",
    "Администрация": "Административные",
    "Финансово-юридические расходы": "Административные",
    "Расходы на офис": "Офис",
    "Бизнес-процессы": "IT и бизнес-процессы",
    "IT": "IT и бизнес-процессы",
    "CRM": "IT и бизнес-процессы",
    "Маркетинг": "Маркетинг",
    "Отдел продаж": "Продажи",
    "Банковские услуги": "Финансовые расходы",
}


_MONTH_HEADER_RE = re.compile(r"^\s*(январь|февраль|март|апрель|май|июнь|июль|август|сентябрь|октябрь|ноябрь|декабрь)\s+20\d{2}\s*$", re.IGNORECASE)


def _discover_month_columns(ws) -> list[tuple[str, int | None, int | None, int | None]]:
    """
    Row 1 has month group labels in merged cells, e.g. «Январь 2026».
    Row 2 has sub-headers («План», «Прогноз», «Факт»).

    Returns list of (month_label, plan_col, forecast_col, actual_col).
    Any of the three cols may be None if missing for that month.
    Ordered by column position (chronological).
    """
    groups = []
    for mr in ws.merged_cells.ranges:
        if mr.min_row == 1 and mr.max_row == 1:
            label = ws.cell(1, mr.min_col).value
            if label and _MONTH_HEADER_RE.match(str(label)):
                groups.append((str(label).strip(), mr.min_col, mr.max_col))
    # Non-merged labels in row 1 (defensive — shouldn't happen if Excel structured properly)
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v and _MONTH_HEADER_RE.match(str(v)) and not any(c0 <= c <= c1 for _, c0, c1 in groups):
            groups.append((str(v).strip(), c, c))
    groups.sort(key=lambda t: t[1])

    result = []
    for label, c_start, c_end in groups:
        plan_col = fore_col = fact_col = None
        for c in range(c_start, c_end + 1):
            sub = ws.cell(2, c).value
            if not sub:
                continue
            s = str(sub).strip().lower()
            if s == "план":
                plan_col = c
            elif s == "прогноз":
                fore_col = c
            elif s == "факт":
                fact_col = c
        result.append((label, plan_col, fore_col, fact_col))
    return result


def parse_overhead() -> pd.DataFrame:
    wb = openpyxl.load_workbook(OVERHEAD_FILE, data_only=True)
    ws = wb["Расходы Шелковый путь"]

    month_cols = _discover_month_columns(ws)
    if not month_cols:
        return pd.DataFrame(columns=["category", "group", "month", "plan", "forecast", "actual"])

    rows = []
    for i in range(3, ws.max_row + 1):
        name = ws.cell(i, 1).value
        if name is None:
            continue
        name = str(name).strip()
        if not name or name in _SKIP_NAMES:
            continue

        # Collect all month values; skip row only if ALL months are None/empty
        month_values = []
        for month, p_col, fo_col, fa_col in month_cols:
            plan = ws.cell(i, p_col).value if p_col else None
            fore = ws.cell(i, fo_col).value if fo_col else None
            fact = ws.cell(i, fa_col).value if fa_col else None
            month_values.append((month, plan, fore, fact))

        if all(p is None and fo is None and fa is None for _, p, fo, fa in month_values):
            continue

        group = _CATEGORY_GROUPS.get(name, "Прочие")

        for month, plan, fore, fact in month_values:
            rows.append({
                "category": name,
                "group": group,
                "month": month,
                "plan": float(plan or 0),
                "forecast": float(fore or 0),
                "actual": float(fact or 0),
            })

    return pd.DataFrame(rows)


def get_overhead_summary(overhead_df: pd.DataFrame, calc_type: str = "actual") -> pd.DataFrame:
    """Aggregate overhead by group and month. calc_type: 'plan'|'forecast'|'actual'."""
    col = calc_type  # 'plan', 'forecast', or 'actual'
    result = (
        overhead_df
        .groupby(["group", "month"])[col]
        .sum()
        .reset_index()
        .rename(columns={col: "amount"})
    )
    return result
