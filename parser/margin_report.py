"""
Parser for margin report files. Files are auto-discovered by pattern:
  "Для ИИ МП NN. <Месяц> <Год> Отчет по марже.xlsx"
(NN = порядковый номер, Месяц = русское название, Год = 4 цифры).

Чтобы добавить новый месяц — положите файл с таким именем рядом с
существующими (в корне проекта) и перезагрузите дашборд.

Sheet "Отчет по марже":
  Row 3 = headers, data from row 4.
  Col 1  (A): комм       — entity code (SF, НЛ, ТЕ)
  Col 2  (B): Юнит
  Col 3  (C): Продукт
  Col 4  (D): Проект     — client / project name
  Col 5  (E): Площадка   — marketplace platform
  Col 7  (G): Открутка с НДС
  Col 8  (H): Открутка без НДС
  Col 11 (K): Работы     — agency revenue from work
  Col 12 (L): Расходы    — direct project expenses
  Col 13 (M): Маржа
  Col 14 (N): Менеджер МП
  Col 15 (O): manager bonus amount
  Col 16 (P): Специалист
  Col 17 (Q): specialist bonus amount
  Col 19 (S): Руководитель напр

Sheet "Отчет по ЗП":
  Row 4 = headers, data from row 5.
  Col 1: Группы, Col 2: Role, Col 3: ФИО, Col 4: ФИКС
  Col 7: manager bonus, Col 8: specialist bonus, Col 9: activity total
  Col 11: прочие доплаты, Col 12: Отпуск/Больничный
  Col 13: Итого начислено, Col 14: выдали 1С
"""

import re
import openpyxl
import pandas as pd
from pathlib import Path

DATA_DIR = Path(__file__).parent.parent

_FILE_RE = re.compile(
    r"Для ИИ МП\s*(?P<idx>\d+)\.\s*(?P<month>\S+)\s+(?P<year>\d{4})\s+Отчет по марже\.xlsx$",
    re.UNICODE | re.IGNORECASE,
)

_MONTH_ORDER = {
    "январь": 1, "февраль": 2, "март": 3, "апрель": 4, "май": 5, "июнь": 6,
    "июль": 7, "август": 8, "сентябрь": 9, "октябрь": 10, "ноябрь": 11, "декабрь": 12,
}


def _discover_files() -> dict[str, Path]:
    """Scan DATA_DIR for margin files, return {month_label: path} sorted chronologically."""
    found = []
    for p in DATA_DIR.glob("Для ИИ МП *.xlsx"):
        m = _FILE_RE.match(p.name)
        if not m:
            continue
        month = m.group("month").strip()
        year = int(m.group("year"))
        label = f"{month} {year}"
        sort_key = (year, _MONTH_ORDER.get(month.lower(), int(m.group("idx"))))
        found.append((sort_key, label, p))
    found.sort(key=lambda t: t[0])
    return {label: path for _, label, path in found}


FILES: dict[str, Path] = _discover_files()

# Row skip marker text that identifies totals / section headers
_TOTAL_MARKERS = {"ИТОГО", "Итого", "итого"}


def _extract_platform(площадка: str | None) -> str:
    if not площадка:
        return "Прочее"
    p = str(площадка)
    for mp in ("Ozon", "Wildberries", "Яндекс.Маркет", "Яндекс Маркет"):
        if mp.lower() in p.lower():
            return mp
    if "KPI" in p or "kpi" in p.lower():
        return "Премия KPI"
    return p


def parse_margin_sheet(month: str) -> pd.DataFrame:
    path = FILES[month]
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Отчет по марже"]

    rows = []
    for i in range(4, ws.max_row + 1):
        def c(col): return ws.cell(i, col).value

        entity = c(1)
        project = c(4)

        # Skip blank rows and total rows
        if entity is None and project is None:
            continue
        if project in _TOTAL_MARKERS or entity in _TOTAL_MARKERS:
            continue

        turnover_vat = c(7) or 0.0
        turnover = c(8) or 0.0
        works = c(11) or 0.0
        expenses = c(12) or 0.0
        margin = c(13) or 0.0

        # Skip rows with no financial data
        if turnover_vat == 0 and works == 0 and margin == 0:
            continue

        rows.append({
            "month": month,
            "entity": entity or "",
            "unit": c(2) or "",
            "product": c(3) or "",
            "project": project or "",
            "platform_raw": c(5) or "",
            "platform": _extract_platform(c(5)),
            "turnover_vat": float(turnover_vat),
            "turnover": float(turnover),
            "works": float(works),
            "expenses": float(expenses),
            "margin": float(margin),
            "manager": c(14) or "",
            "manager_bonus": float(c(15) or 0),
            "specialist": c(16) or "",
            "specialist_bonus": float(c(17) or 0),
            "director": c(19) or "",
        })

    return pd.DataFrame(rows)


def parse_salary_sheet(month: str) -> pd.DataFrame:
    path = FILES[month]
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Отчет по ЗП"]

    rows = []
    for i in range(5, ws.max_row + 1):
        def c(col): return ws.cell(i, col).value

        name = c(3)
        fiks = c(4)

        # Skip blank rows and total rows (no name or no fixed salary)
        if name is None or fiks is None:
            continue
        if not str(name).startswith("ФИО"):
            continue

        rows.append({
            "month": month,
            "group": c(1) or "",
            "role": c(2) or "",
            "name": str(name),
            "fiks": float(fiks or 0),
            "manager_bonus": float(c(7) or 0),
            "specialist_bonus": float(c(8) or 0),
            "activity": float(c(9) or 0),
            "other_pay": float(c(11) or 0),
            "vacation_sick": float(c(12) or 0),
            "total_accrued": float(c(13) or 0),
            "paid_1c": float(c(14) or 0),
        })

    df = pd.DataFrame(rows)
    return df


def load_all_months() -> tuple[pd.DataFrame, pd.DataFrame]:
    """Returns (margin_df, salary_df) for all available months."""
    margin_frames, salary_frames = [], []
    for month in FILES:
        if FILES[month].exists():
            margin_frames.append(parse_margin_sheet(month))
            salary_frames.append(parse_salary_sheet(month))

    margin_df = pd.concat(margin_frames, ignore_index=True) if margin_frames else pd.DataFrame()
    salary_df = pd.concat(salary_frames, ignore_index=True) if salary_frames else pd.DataFrame()
    return margin_df, salary_df
